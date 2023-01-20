import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, date
from openpyxl import Workbook

# Esto libreria nos ayuda a crear pivot tables
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

######################
#Tablas e imagenes
# Primero importo algunas funciones para tablas e imagenes
from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl.chart import PieChart, Reference, Series,PieChart3D,LineChart, BarChart


class Format:

    def set_table():
        df = pd.read_excel('docs/temp/table_test.xlsx')

        # Aqui buscamos baer el tamaño de la table_test
        # Agregamos 1 para que cubra el ultimo row de abajo
        # Lo pasamos a string para poder pegarlo en una cadena
        size = str(1+(df.shape[0]))

        # Primero abro el libro que ya tiene la columna incluida
        wb = load_workbook('docs/temp/table_test.xlsx')
        #----------------------#
        #       WORK SHEET     #
        #----------------------#
        # grab the active worksheet
        # This will create the active sheet on this work BOOK
        ws = wb.active
        #Provide le sheet with a name
        ws.title = "Table"
        # Creamos un objeto de data que guarde los datos de la tabla
        # el parametro 'ref' indica el rango de los datos
        tab = Table(displayName='Table1',ref = 'B1:Q'+size)
        #print('A1:'+'AE'+t_size)
        # Ahora le damos estilo a la tabla
        # Creo una variable que guarde el objeto de estilo
        style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False, showLastColumn= False,
                                                            showRowStripes=True, showColumnStripes=True)
        # ahora asigno la variable
        tab.tableStyleInfo = style

        # Por ultimo agregamo la tabla al worksheet
        ws.add_table(tab)

        wb.save('docs/temp/final_table.xlsx')

    def set_graphic():

        # Primero abro el libro que ya tiene la columna incluida
        wb = load_workbook('docs/temp/final_table.xlsx')
        ws = wb.active

        df = pd.read_excel('docs/temp/final_table.xlsx')

        size = df.shape[0]
        print(size)
        #-------------------------------------
        # Creamos otro tab pra guardar el grafico
        # With teh work book object we call the create fucntion
        ws_1 = wb.create_sheet("Statistics")

        #-------------------------------------
        # CREAMOS grafico
        #-------------------------------------

        # Ahora que tenemos informacion tenemos que decidir como se presenta en la grafica
        chart = BarChart()

        # Especificamos desde donde, hasta donde queremos la data
        # Tenemos que encontrar el tamaño de la tabla
        labels = Reference(ws,min_col=2, min_row=1, max_row=size)

        data = Reference(ws,min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row)


        chart.add_data(data,titles_from_data=True)
        chart.set_categories(labels)
        chart.title = 'Ice crema Flavor'

        # Añadimos el gráfico a la pestaña correcta
        # Decimos en que celda deseo que comience.
        ws_1.add_chart(chart,'Q5')

        # Cunado tenemos todo lo que queremos guardamos el archivo
        wb.save('docs/temp/graphs.xlsx')

        return wb

    def new_column():
        ##################
        #  Add columns
        ##################
        # Ahora vamos a gruardar el reporte combinado a un excel
        wb = load_workbook('docs/1_jorunals.xlsx')
        ws = wb.active
        #Provide le sheet with a name
        ws.title = "Data"

        # Ahora creamos la cell que queremos agregar
        # Hay que tener en cuant el tamaño del sheet para poder ubicarla bien en la columna correcta
        # Creo una variable donde guardo la cell

        test = ws['AE1']
        # test.font = Font(bold=True)
        test.value = 'Month-Year'

        #################
        #  FIND SIZE DF
        #################
        # Podemos averiguar el tamaño de la coumna facilment con pandas
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        df = pd.read_excel('docs/1_jorunals.xlsx')
        size = df.shape[0]
        ##########################
        #  FOR LOOP NEW COLUMN
        ##########################
        # Ahora creamos un forloop para interactuar con todos los rows en estas columnas
        # RECORDAR que no debemos incluir el primer row porque son HEADING y no values
        # Ponemos de limite la variable 'df'
        # Le ponemos el size mas 2 porque me fglatan dos espacio
        # Le decimos que comience desde 2 pero el DF me cuenta solo los spacios con data y los los encabezados
        # Para compensar esto hacemos el truco
        for row in range(2,(size+2)):
            # En cada interaccion pillamos el numero de row para el MES y el año
            # Lo convertimos a STR para poder hacerlo una cadena de texto
            m = str(ws[f'AC{row}'].value)
            y = str(ws[f'AD{row}'].value)

            # Creamos la cadena de texto con la fecha complete
            a_date = "1"+"/"+m+"/"+y

            # En el último parametro especifico el formato de fecha que deseamos tener '%B %Y'
            # B% da el mes con nombre completo
            # m% nos da el numero del mes

            #############################
            #       NEW columna
            #############################
            # En la columna AE ya le hemos asignado nombre
            # Ahora necesitamos asignarle los valores
            ws[f'AE{row}'] = datetime.strptime(a_date, "%d/%m/%Y").strftime('%B %Y')

        #SAlvamos el doc como excel
        wb.save('docs/temp/new_column.xlsx')

    def convert_excel():
        #################
        #  FIND SIZE DF
        #################
        # Podemos averiguar el tamaño de la coumna facilment con pandas
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        df = pd.read_excel('docs/temp/new_column.xlsx')

        # select  columns to display
        df = df[['Month-Year',
                'Unique_Item_Requests',
                'Total_Item_Requests',
                'Platform',
                'Subject',
                'OrderDescription',
                'OrderNumber',
                'OrderExpired',
                'OwnedByCustomerNumber',
                'OwnedByCustomer',
                'UsedByCustomerNumber',
                'UsedByCustomer',
                'Group',
                'User',
	            'Month',
                'Year']]

        # saving the excel
        # df.to_excel('docs/temp/table_test.xlsx',index=False, header=False)
        df.to_excel('docs/temp/table_test.xlsx')

    def set_pivot(wb):

        #-------------------------------------
        # Primero leemos el dataframe_to_rows
        # -------------------------------------
        #read in data from relevant excel file
        #df = pd.read_excel('docs/temp/graphs.xlsx',index_col='Date',parse_dates=True)
        df = pd.read_excel('docs/temp/graphs.xlsx')
        #----------------------------------------------------------------------------------------
        #-------------------------------------
        # Trabajamos el dataframe con funcion nativa de Oopenpyxl
        # -------------------------------------
        #this fucntion is a native openpyxl function that allows us towork specifically with Pandas DataFrames
        #it allows us to iterate through the rows and append each one to our active worksheet.

        for r in dataframe_to_rows(df, index=True, header=True):
                ws_1.append(r)

        #we could just save the file at this point and we would have a full working ".xlsx" version of the data
        #however it would suffer fro, the same column width issues that we experinced when using Pandas to save.
        #instead we iterate through each column in the worksheet and set the width variable to the maximum width
        #of data held witihin that particular column, then set the width of that column to that value, plus 2
        #(just to make sure the data shows fully)

        for column_cells in ws_1.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            ws_1.column_dimensions[column_cells[0].column].width = length + 2
            #ws_1.column_dimensions[column_cells[0].column].width = 4

        #set title of "Date" column
        ws_1['A1'] = "Date"
        #create pivoyt table using Pandas

        # Values--------> Es lo que va a rellenar los espacios
        # Index ---------> Es lo que va a servir como indice vertical
        # Columns -----------> Son las columnas que se despliegan

        data_piv = df.pivot_table(values=['Unique_Item_Requests','Total_Item_Requests'],index='Platform',columns='Year',aggfunc='sum')

        #create pivoyt table using Pandas

        for r in dataframe_to_rows(data_piv, index=True):
                ws_1.append(r)

        #-------------------------------------
        # CREAMOS grafico
        #-------------------------------------

        # Ahora que tenemos informacion tenemos que decidir como se presenta en la grafica
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Usage Statistics"

        chart.y_axis.title = 'Units Sold'
        chart.x_axis.title = 'Platform'

        data = Reference(ws_1,min_col=2, min_row=1, max_row=ws_1.max_row, max_col=ws_1.max_column)

        # Especificamos desde donde, hasta donde queremos la data
        # Tenemos que encontrar el tamaño de la tabla
        labels = Reference(ws_1,min_col=2, min_row=2, max_row=ws_1.max_row)


        chart.add_data(data,titles_from_data=True)
        chart.set_categories(labels)
        chart.shape = 4

        #place chart at specific cell reference - one row after the last currently populated row
        # ws_1.add_chart(chart, "A"+str(ws_1.max_row +  1))
        ws_1.add_chart(chart, "I3")

        wb.save('docs/temp/report.xlsx')
