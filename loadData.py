import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

from openpyxl.chart import BarChart, Reference

import datetime

#---------------------#
#   LOAD WORK BOOK    #
#---------------------#

wb = load_workbook("docs/test.xlsx")

sheet = wb.active

x1 = sheet['A1']
x2 = sheet['A2']
#using cell() function
x3 = sheet.cell(row=3, column=1)

print("The first cell value:",x1.value)
print("The second cell value:",x2.value)
print("The third cell value:",x3.value)

#Getting list of all sheet available in workbook
print(wb.get_sheet_names())

#---------------------#
#      CREATE CHART   #
#---------------------#
chart = BarChart()
values = Reference(worksheet=sheet,
                 min_row=1,
                 max_row=8,
                 min_col=2,
                 max_col=3)

chart.add_data(values, titles_from_data=True)
sheet.add_chart(chart, "T2")

wb.save("docs/student_chart.xlsx")
