import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
#---------------------#
#       WORK BOOK     #
#---------------------#
wb = Workbook()

#----------------------#
#       WORK SHEET     #
#----------------------#
# grab the active worksheet
# This will create the active sheet on this work BOOK
ws = wb.active
#Provide le sheet with a name
ws.title = "Titulo que queramos"
# With teh work book object we call the create fucntion
ws_1 = wb.create_sheet("Test1")
#       ACTIVE     #
# 0 Means that it goes before anything else
# THis doesnt mean is the active one
# Just means will be first
ws_2 = wb.create_sheet("Test2",0)


# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
