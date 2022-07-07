from openpyxl import load_workbook

def iterating_over_values(path, sheet_name):
    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f"'{sheet_name}' not found. Quitting.")
        return
    sheet = workbook[sheet_name]

    """
        min_col (int) – smallest column index (1-based index)
        min_row (int) – smallest row index (1-based index)
        max_col (int) – largest column index (1-based index)
        max_row (int) – largest row index (1-based index)
        values_only (bool) – whether only cell values should be returned
    """
    # We can set thi va discovering the size of the worksheet
    for value in sheet.iter_rows(
        min_row=2, max_row=137, min_col=1, max_col=10,
        values_only=True):
        print(value)

if __name__ == "__main__":
    iterating_over_values("test.xlsx", sheet_name="Journals")
