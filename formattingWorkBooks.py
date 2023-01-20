import openpyxl
# Importamos las funciones de esta forma
# Optenemos todo el poder de la libreria para dar formato profesiona a worksheets.
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl import Workbook
from openpyxl import load_workbook

# Ahora creamos el Workbook
wb = Workbook()
ws = wb.active

# Ahora lleno el excel con algo de data
for i in range(1,20):
    ws.append(range(300))

# MERGE CELLS FUCTIONS
# YOu need to specify the range of cells
# We pass the slice to the functions as argument
# Marcamos desde la celda que comienza hasta donde termina
# Lo mismo que hariamos con el mouse
ws.merge_cells("B2:E5")
#ws.unmerge.cells()
# Tambien lo podemos hacer pasando - desde donde - hasta donde - queremos hacer el MERGE
# Por ejemplo queremos merge desde B2 hasta el E5
# They are addressed form the top left cell
#ws.merge_cells(start_row= 2, start_column= 2, end_row=5, end_column=5)

# Ahora podemos convertir todo el MERGE en una variable
# Tomamos como referencia cual va a ser el top left corner de la seleccion
# Conicide con la celda donde comenzamos nuestra selseccion.
cell = ws['B2']

# Damos propiedades al texto de la variable CELL
#-------> Not working------
#cell.font = Font(color= colors.RED, size=20, italic=True)

# Aqui escribimos el texto que va a tener la variable
cell.value = 'Merged Cell'
# POdemos ubicar el texto donde queramos
# ALIGNMENT toma dos posiciones : Horizontal y vertical
cell.alignment = Alignment(horizontal= 'right', vertical='bottom')

# Con relacion al color lo podemos trabajar de forma porfesional
# Podemos dar un relleno solido o una transicion suave con GRADIENT PatternFill
# Especificamos el color donde comenzamos y el color donde se termina
# HEX decimal color en RRGGBB format
# Esto significa que los primero 2 valores= RED, segundos 2 = Green, terceros 2= BLUE
cell.fill = GradientFill(stop=("000000","FFFFFF"))

wb.save('docs/formatting.xlsx')


##############################
#     Name styles
##############################

# SOn nombres que especificamos para que guarde los stilos y no tenerlos que escribir cismpre que los necesitamos.
# Primero creamos la variable
hightlight = NamedStyle(name ='hightlight')
hightlight.font = Font(bold=True)
 # Ahora creamos una variable para que los bordes sean anchos
bd = Side(style='thick', color='000000')
hightlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

 # Ahora tememos que colorear lo de adentro con un color . Amarillo por ejemplo
 # En lugar de utilizar el GRADIENT FILL vamosa  utilizar el PatternFill
# Si quito el blue de la ecuacion nos da un amarillo
hightlight.fill = PatternFill('solid', fgColor='FFFF00')

 # Ahora si quiero aplicar este estilo a una celdas en particular tengo que indicar el punto de partida.
 # Inicio un contador de las celdas que llevaran este estilo
count = 0

# Para recorrer las celdas existe una funcion de FOR Do que viene con la libreria
# Primero seleccionamos los minimos y despues los maximos para que no se salga de control
# Esto nos pinta un tipo de escalera
for col in ws.iter_cols(min_col=8, min_row=1, max_col=30, max_row=30):
    #Remeber that en la funcion ITER_COLS, columns store a list
    # Por eso podemos utilizar el couter para indexar la lista paralos rows we want
    # Cada vuelta del cauter nos movemos un row para abajo
    col[count].style = hightlight
    count = count + 1
wb.save('docs/formatting.xlsx')
