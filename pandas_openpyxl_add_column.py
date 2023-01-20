from datetime import datetime, date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Esto es para poder utilizar formulas
from openpyxl.utils import FORMULAE

# Ayuda a los estilo de la celdas
from openpyxl.styles import NamedStyle


##################
#  Add columns
##################
# Ahora vamos a gruardar el reporte combinado a un excel
wb = load_workbook('docs/stats/all_docs.xlsx')
ws = wb.active
ws.title = "Data"
#  Add columns
##################
# Ahora creamos la cell que queremos agregar
# Hay que tener en cuant el tamaño del sheet para poder ubicarla bien en la columna correcta
# Creo una variable donde guardo la cell

test = ws['AE1']
test.font = Font(bold=True)
test.value = 'Month-Year'

# Para poder replicar la formula para agregar fecha con meses
# Necesito crear variable que me guarden los valores con lo que voy a operar

#month, year = ['AC','AD']


# Podemos averiguar el tamaño de la coumna facilment con pandas
# Llamo a un dataframe para buscar la info rapido y asigno una variable
df = pd.read_excel('docs/stats/all_docs.xlsx')
size = df.shape[0]
# Ahora creamos un forloop para interactuar con todos los rows en estas columnas
# RECORDAR que no debemos incluir el primer row porque son HEADING y no values
# Ponemos de limite la variable 'df'
# Le ponemos el size mas 2 porque me fglatan dos espacio
# Le decimos que comience desde 2 pero el DF me cuenta solo los spacios con data y los los encabezados
# Para compensar esto hacemos el truco

# now = datetime.now()


for row in range(2,(size+2)):
    # En cada interaccion pillamos el numero de row para el MES y el año
    # Lo convertimos a STR para poder hacerlo una cadena de texto
    m = str(ws[f'AC{row}'].value)
    y = str(ws[f'AD{row}'].value)

    # Creamos la cadena de texto con la fecha complete
    a_date = "1"+"/"+m+"/"+y

    # En el último parametro especifico el formato de fecha que deseamos tener '%B %Y'
    # B% da el mes con nombre completo
    # m% nos da el numero del mes

    #############################
    #       NEW columna
    #############################
    # En la columna AE ya le hemos asignado nombre
    # Ahora necesitamos asignarle los valores
    ws[f'AE{row}'] = datetime.strptime(a_date, "%d/%m/%Y").strftime('%B %Y')

wb.save('docs/stats/all_docs_plus_column.xlsx')
