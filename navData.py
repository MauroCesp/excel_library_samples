import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

from openpyxl.chart import BarChart, Reference
#---------------------#
#   LOAD WORK BOOK    #
#---------------------#
def open_workbook(path):
    workbook = load_workbook(filename=path)
    print(f"Worksheet names: {workbook.sheetnames}")
    print(workbook.sheetnames[1])
    sheet = workbook.active
    print(f"The title of the active Worksheet is: {sheet.title}")


if __name__ == "__main__":
    # Execute the function to load the info
    open_workbook("test.xlsx")
